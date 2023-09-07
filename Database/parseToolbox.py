# BGlas - InfosecDad
# This is some hacky python to parse a SAMM 2.0.x excel toolbox 
# and generate the SQL insert statements to populate the SAMM benchmark DB

from openpyxl import load_workbook
import os

# I'm sure there are better ways to do this, just short on time
# so this is the mapping of question codes to UIDs in the YAML model
# because the YAML UIDs are never used in the questionnaire, the code are
questions = {
'D-SA-A-1-1':'c4eb5618d1814173a995f8aea96f1c0b',
'D-SA-B-1-1':'47c8fb0cae5944d090d7f73f7632dc9f',
'D-SA-A-2-1':'3f1a3a84c85f4e339bcc5c9ecca5c73a',
'D-SA-B-2-1':'b63b5fa0f5bc455bb5b1dd9168c44000',
'D-SA-A-3-1':'cb88049632b54a15b3d610c4d492e83e',
'D-SA-B-3-1':'f4722a4fdfc44a45be5b5ee8dd7b735f',
'D-SR-A-1-1':'645207bf33584cc6a535e42bae7667c9',
'D-SR-B-1-1':'795e7ddd03f2443c851e34fc6e023d71',
'D-SR-A-2-1':'2d458a65858c48af94f25f9858bd8ed7',
'D-SR-B-2-1':'dffdd9659e6243d7bcbcbc0dff4429fc',
'D-SR-A-3-1':'dad421c501994b0fa2b2ab94ffe61176',
'D-SR-B-3-1':'497753e656514aa6bdf3030bebcb3fbe',
'D-TA-A-1-1':'71c02652a9ba4b10a0cc0179c7ce869f',
'D-TA-B-1-1':'e9dcf4f79e2e487fb74df5e15a14a91b',
'D-TA-A-2-1':'897306b66f16454eab7b5a2355d31c11',
'O-IM-B-1-1':'0d889a913d484eb39b80f096f3a66019',
'D-TA-B-2-1':'42cfabd13db34fd0b35e92af917eb1b8',
'D-TA-A-3-1':'7e541611f3c749f285ac27f0a9ba7d55',
'D-TA-B-3-1':'cd639e5458ca4f60be60bc9d47314648',
'G-EG-A-1-1':'3d801ea0b0ad4c44ba2b0408ebcc750b',
'G-EG-B-1-1':'21a9b65765a844e0b27a074f2b4306a1',
'G-EG-A-2-1':'1962ef9fe4cf488a8d10ccbcdc8bb926',
'G-EG-B-2-1':'fe0485b5026d4b2b9a7c99260addc912',
'G-EG-A-3-1':'c7147e96d99849a994e63d5732c26220',
'G-EG-B-3-1':'871a30e6aaef4905a3d3302ceea808f4',
'G-PC-A-1-1':'d15bfcd426d24a43b9417a0464b3af44',
'G-PC-B-1-1':'4bb7dd93f5874730bd84b41cb56ce60a',
'G-PC-A-2-1':'54aca12ea18e4073becbdd356cd3b3ef',
'G-PC-B-2-1':'84d5a7f8d9e647db95288329f64fc41d',
'G-PC-A-3-1':'073b78f0ce314757a05b15d9c6f96703',
'G-PC-B-3-1':'6a81ec4bcd1f43de95b91a7f50a40244',
'G-SM-A-1-1':'64f49a28334e4a40a04e534225a941d2',
'G-SM-B-1-1':'6aa85d5f3c03428aa064447fa50fa66b',
'G-SM-A-2-1':'8fd0374f0b63476eacb3cadca99b1538',
'G-SM-B-2-1':'26f12b057f2b44f386d9af391383b590',
'G-SM-A-3-1':'74eaee0cbf454a46adeb2619850bbcb3',
'G-SM-B-3-1':'efdc34fbb75b405b8107d63d58fa7286',
'I-DM-A-1-1':'7f92b3f940cb4229a6079016269d76c6',
'I-DM-B-1-1':'29df0959af8f403383c8ad01a0f3c478',
'I-DM-A-2-1':'5b333ff2dd474381b0af595ff13dcdcf',
'I-DM-B-2-1':'1a849e8fd3ae41a4b3675947482426da',
'I-DM-A-3-1':'6b5eac7b9e2f49e2a2cda600ef70ad99',
'I-DM-B-3-1':'0247b5b573b843609fdff791a1cc7c1b',
'I-SB-A-1-1':'70d6044a223b402c8e2b6f9d1e936641',
'O-IM-A-1-1':'1aa1ec270db6496d8229e094fdbbc6e5',
'I-SB-B-1-1':'1e28b82cc3ba4e4ea2552746e17c25af',
'I-SB-A-2-1':'b5d33583538b4878bb4674a5f838b8ea',
'I-SB-B-2-1':'ef798e60155d453186364c94a8f8935d',
'I-SB-A-3-1':'ee775955bf7f48d294c75f6384232f48',
'I-SB-B-3-1':'418e98e2939546e69a24d0c3c4c8d217',
'I-SD-A-1-1':'24697d43707b4d83a6a5819a9db9a75d',
'I-SD-B-1-1':'37c6a5618a6344b386eb872619cfe53f',
'I-SD-A-2-1':'7ef6753cda0d4da6ad194e56650f584d',
'I-SD-B-2-1':'c5f66e97db174d8c9dc2d82fbad9d4e3',
'I-SD-A-3-1':'f87fddbf283a4c38b45a3125d827dd27',
'I-SD-B-3-1':'50af7a14ddf2408fb4576e3972cf13e3',
'O-EM-A-1-1':'1e005e11997f4929a12fdb939599e77e',
'O-EM-B-1-1':'180e194b165d421c9d2c89258195a792',
'O-EM-A-2-1':'41d33402a94c49538554ce77e9de6a72',
'O-EM-B-2-1':'0844b1a3be8b49ec83c7377a9f797cfc',
'O-EM-A-3-1':'f4ec030280ee417099eaf12752a542ae',
'O-EM-B-3-1':'6e72179a31c04024bb649346bfb03eb5',
'O-IM-A-2-1':'13b9816c06444ba99584e657bfa5833d',
'O-IM-B-2-1':'dbb83b0d5b504db6a170710df4df347f',
'O-IM-A-3-1':'09744a244f8d4076bec35130da92ea2b',
'O-IM-B-3-1':'91bd2bdc1c734d8dbffc30e37158ab00',
'O-OM-A-1-1':'41b0c2ab37774000b2b81077605bbd93',
'O-OM-B-1-1':'b50d81aba3734cb59d618fd74bc5c99e',
'O-OM-A-2-1':'424c004afec041058e89c396c9e88930',
'O-OM-B-2-1':'f9c5a8649ddf4168b6f62d0018a32704',
'O-OM-A-3-1':'8176c4588bdd4e979e3c82566450696a',
'O-OM-B-3-1':'54ad3e4182844623b3290901c1a1932d',
'V-AA-A-1-1':'2da7acf355814b75ab971fec36048f11',
'V-AA-B-1-1':'b920062a62d84883af0c167955ec10de',
'V-AA-A-2-1':'b133d28653bc47a8a8574d1c60ec34f0',
'V-AA-B-2-1':'2784c6272d174dcf932b188a69a8917d',
'V-AA-A-3-1':'ba213b2d5fc844a386010ca53cf87fb8',
'V-AA-B-3-1':'df317b6ac5de4815a1ba7ac558d0263e',
'V-RT-A-1-1':'cb085e0a25724700bc10c73cfcc8f6a5',
'V-RT-B-1-1':'749893a53df24c32bc887b6d5f7b3f7b',
'V-RT-A-2-1':'d6b26a63243d4142bd2a8317e2875a03',
'V-RT-B-2-1':'b2afb33fe6ef4b6e90e029059f7a7124',
'V-RT-A-3-1':'35413be30d9f415dbde45edbe3b17f31',
'V-RT-B-3-1':'814caca2fc5241dcb90c48302ac031b2',
'V-ST-A-1-1':'bef645da8ccd477bbd10685dd52ad40e',
'V-ST-B-1-1':'77dd81adf35f43608408e548c4972136',
'V-ST-A-2-1':'c013b6f9d973425cb63f21f4f8b84c30',
'V-ST-B-2-1':'9a2af155ba424edfb321aa7592a09ed5',
'V-ST-A-3-1':'009a8fafe5dd41889947a6b2c6769bbe',
'V-ST-B-3-1':'b73bf8f0462340659e252e6471c6e831'
}

# Answer mapping because it's not aligned...
# 'Toolbox':'File'
# Since they are jumbled on the creation of the toolbox
# I need to map from the toolbox letter to the YAML model letter
ftoolboxmap = {
'F':'A',
'C':'B',
'A':'C',
'I':'D',
'H':'E',
'L':'F',
'M':'G',
'R':'H',
'Q':'I',
'O':'J',
'P':'K',
'S':'L',
'U':'M',
'Y':'N',
'V':'O',
'N':'P',
'K':'Q',
'G':'R',
'J':'S',
'B':'U',
'E':'V',
'D':'W',
'W':'X',
'X':'Y',
'T':'Z',
}

# The YAML model doesn't have UIDs for each answer, only the overall answerset
# I've added (0-3) suffix to the UIDs for the four possible answers
# And then mapped them in here.
answers = [
['A',0, 'f77bd45a28c8493dbba6e53b2eafa20f0'],
['A',0.25, 'f77bd45a28c8493dbba6e53b2eafa20f1'],
['A',0.5, 'f77bd45a28c8493dbba6e53b2eafa20f2'],
['A',1, 'f77bd45a28c8493dbba6e53b2eafa20f3'],
['B',0, '8c89e8daf71d425abaca53edc01f6afa0'],
['B',0.25, '8c89e8daf71d425abaca53edc01f6afa1'],
['B',0.5, '8c89e8daf71d425abaca53edc01f6afa2'],
['B',1, '8c89e8daf71d425abaca53edc01f6afa3'],
['C',0, '9a87d689fe35441aabf1ad4b7048b61e0'],
['C',0.25, '9a87d689fe35441aabf1ad4b7048b61e1'],
['C',0.5, '9a87d689fe35441aabf1ad4b7048b61e2'],
['C',1, '9a87d689fe35441aabf1ad4b7048b61e3'],
['D',0, 'f96770095fab4afbb27949c2242e47c20'],
['D',0.25, 'f96770095fab4afbb27949c2242e47c21'],
['D',0.5, 'f96770095fab4afbb27949c2242e47c22'],
['D',1, 'f96770095fab4afbb27949c2242e47c23'],
['E',0, 'd096060a4d864133afcbdd1397b958270'],
['E',0.25, 'd096060a4d864133afcbdd1397b958271'],
['E',0.5, 'd096060a4d864133afcbdd1397b958272'],
['E',1, 'd096060a4d864133afcbdd1397b958273'],
['F',0, '3d4c5c80278b4a58b80d5590858044460'],
['F',0.25, '3d4c5c80278b4a58b80d5590858044461'],
['F',0.5, '3d4c5c80278b4a58b80d5590858044462'],
['F',1, '3d4c5c80278b4a58b80d5590858044463'],
['G',0, '612bf4ec249f4e9d86f9e36dbf5118210'],
['G',0.25, '612bf4ec249f4e9d86f9e36dbf5118211'],
['G',0.5, '612bf4ec249f4e9d86f9e36dbf5118212'],
['G',1, '612bf4ec249f4e9d86f9e36dbf5118213'],
['H',0, '381e1e37a19c488ab045a8a5125521410'],
['H',0.25, '381e1e37a19c488ab045a8a5125521411'],
['H',0.5, '381e1e37a19c488ab045a8a5125521412'],
['H',1, '381e1e37a19c488ab045a8a5125521413'],
['I',0, 'e5a12ab46e4645a9ab22aa5a1ebe562f0'],
['I',0.25, 'e5a12ab46e4645a9ab22aa5a1ebe562f1'],
['I',0.5, 'e5a12ab46e4645a9ab22aa5a1ebe562f2'],
['I',1, 'e5a12ab46e4645a9ab22aa5a1ebe562f3'],
['J',0, '6c3e82e127264b92b25b732d85286d720'],
['J',0.25, '6c3e82e127264b92b25b732d85286d721'],
['J',0.5, '6c3e82e127264b92b25b732d85286d722'],
['J',1, '6c3e82e127264b92b25b732d85286d723'],
['K',0, '14ad9a12e44f4079abc610010292f35e0'],
['K',0.25, '14ad9a12e44f4079abc610010292f35e1'],
['K',0.5, '14ad9a12e44f4079abc610010292f35e2'],
['K',1, '14ad9a12e44f4079abc610010292f35e3'],
['L',0, 'c1d15e1f5c8946d381f508db29b264730'],
['L',0.25, 'c1d15e1f5c8946d381f508db29b264731'],
['L',0.5, 'c1d15e1f5c8946d381f508db29b264732'],
['L',1, 'c1d15e1f5c8946d381f508db29b264733'],
['M',0, 'b6fd4b86ecf04955befe9322ff338ca80'],
['M',0.25, 'b6fd4b86ecf04955befe9322ff338ca81'],
['M',0.5, 'b6fd4b86ecf04955befe9322ff338ca82'],
['M',1, 'b6fd4b86ecf04955befe9322ff338ca83'],
['N',0, 'f678b7a00f2441148087d48f8e0a6ad10'],
['N',0.25, 'f678b7a00f2441148087d48f8e0a6ad11'],
['N',0.5, 'f678b7a00f2441148087d48f8e0a6ad12'],
['N',1, 'f678b7a00f2441148087d48f8e0a6ad13'],
['O',0, '66e3e11eb8404fb6880377e5396096780'],
['O',0.25, '66e3e11eb8404fb6880377e5396096781'],
['O',0.5, '66e3e11eb8404fb6880377e5396096782'],
['O',1, '66e3e11eb8404fb6880377e5396096783'],
['P',0, '01b2ac64461d4ec6b40843a4c77e1ba60'],
['P',0.25, '01b2ac64461d4ec6b40843a4c77e1ba61'],
['P',0.5, '01b2ac64461d4ec6b40843a4c77e1ba62'],
['P',1, '01b2ac64461d4ec6b40843a4c77e1ba63'],
['Q',0, '608f87d59da44e589f0090790675ed230'],
['Q',0.25, '608f87d59da44e589f0090790675ed231'],
['Q',0.5, '608f87d59da44e589f0090790675ed232'],
['Q',1, '608f87d59da44e589f0090790675ed233'],
['R',0, 'f3534ade73d8469e879c74b4e0a4eb3d0'],
['R',0.25, 'f3534ade73d8469e879c74b4e0a4eb3d1'],
['R',0.5, 'f3534ade73d8469e879c74b4e0a4eb3d2'],
['R',1, 'f3534ade73d8469e879c74b4e0a4eb3d3'],
['S',0, 'e11943dd1978471486d5db4b04e3047d0'],
['S',0.25, 'e11943dd1978471486d5db4b04e3047d1'],
['S',0.5, 'e11943dd1978471486d5db4b04e3047d2'],
['S',1, 'e11943dd1978471486d5db4b04e3047d3'],
['U',0, '439e7b91e6b446ae83b4d1efe831a97d0'],
['U',0.25, '439e7b91e6b446ae83b4d1efe831a97d1'],
['U',0.5, '439e7b91e6b446ae83b4d1efe831a97d2'],
['U',1, '439e7b91e6b446ae83b4d1efe831a97d3'],
['V',0, 'e0fcc49a200847eab218c04e2c80490a0'],
['V',0.25, 'e0fcc49a200847eab218c04e2c80490a1'],
['V',0.5, 'e0fcc49a200847eab218c04e2c80490a2'],
['V',1, 'e0fcc49a200847eab218c04e2c80490a3'],
['W',0, 'f5042ff6c8d44068a9ac3e1bd83497600'],
['W',0.25, 'f5042ff6c8d44068a9ac3e1bd83497601'],
['W',0.5, 'f5042ff6c8d44068a9ac3e1bd83497602'],
['W',1, 'f5042ff6c8d44068a9ac3e1bd83497603'],
['X',0, 'a0d515d66004425e8039cf4197fce2710'],
['X',0.25, 'a0d515d66004425e8039cf4197fce2711'],
['X',0.5, 'a0d515d66004425e8039cf4197fce2712'],
['X',1, 'a0d515d66004425e8039cf4197fce2713'],
['Y',0, 'f0ccf7b66c0a484aa8374a387438bc980'],
['Y',0.25, 'f0ccf7b66c0a484aa8374a387438bc981'],
['Y',0.5, 'f0ccf7b66c0a484aa8374a387438bc982'],
['Y',1, 'f0ccf7b66c0a484aa8374a387438bc983'],
['Z',0, '51466c3df15b45119e3fc68293f160340'],
['Z',0.25, '51466c3df15b45119e3fc68293f160341'],
['Z',0.5, '51466c3df15b45119e3fc68293f160342'],
['Z',1, '51466c3df15b45119e3fc68293f160343']
]

wb = load_workbook(r'C:\Users\BrianGlas\Projects\SAMM\Assessments\RABET-V_SAMM_Assessment_Toolbox_v2.0-VRSystems.xlsx', read_only='true', data_only='true')

# Map the two tabs needed from the excel toolbox
ws = wb['Interview']
sc = wb['Scorecard']

# Update to match the next sequence number for each of these
assess_id = 10
assessor_id = 5
organization_id = 8
method_id = 2
org_name = ws['D10'].value
assess_date = ws['D12'].value
assessor_name = ws['D13'].value

# Scores for Business Functions and Overall
gov = sc['J14'].value
des = sc['J15'].value
imp = sc['J16'].value
ver = sc['J17'].value
ops = sc['J18'].value
avg = (gov + des + imp + ver + ops) / 5

# Scores for Security Practices
sm = sc['C14'].value
pc = sc['C15'].value
eg = sc['C16'].value
ta = sc['C17'].value
sr = sc['C18'].value
sa = sc['C19'].value
sb = sc['C20'].value
sd = sc['C21'].value
dm = sc['C22'].value
aa = sc['C23'].value
rt = sc['C24'].value
st = sc['C25'].value
im = sc['C26'].value
em = sc['C27'].value
om = sc['C28'].value

# INSERT Statements for the high level information about the assessment
insert = "INSERT INTO ORGANIZATION(industry_id, region_id, employee_num, developer_num, appsec_num, public) VALUES (3,4,1,1,0,'{}');".format(org_name)
f = open("Database/insert_assessment.pgsql", "w")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSOR(name, org, description) VALUES ('{}', '{}', '');".format(assessor_name, org_name)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSMENT(sammmodel_id,assessment_id,assessment_date,organization_id,assessor_id,method_id,maturity_score) VALUES (4,{},'{}',{},{},{},{});".format(assess_id, assess_date, organization_id, assessor_id, method_id, avg)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSORORGANIZATION VALUES ({},{});".format(assessor_id, organization_id)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

# INSERT INTO ASSESS BUSINESS FUNCTION SCORES
insert = "INSERT INTO ASSESSFUNCTION VALUES ({}, '102ad02df5dc4a8eb3837ef4ca2c1af4','{}');".format(assess_id, gov)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSFUNCTION VALUES ({}, '88c296acaae841a2b2fc5314bff44cb4','{}');".format(assess_id, des)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSFUNCTION VALUES ({}, '8aa8154b83434e73b3ca8c0e9b654417','{}');".format(assess_id, imp)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSFUNCTION VALUES ({}, 'fa340fa1816244d79f369ae82e998368','{}');".format(assess_id, ver)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSFUNCTION VALUES ({}, '942d679b0c9e41909f8bde728fdb1259','{}');".format(assess_id, ops)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

# INSERT INTO ASSESS PRACTICE SCORES
insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '32b3bdd85d3a4d53827960004f9d1c7e','{}');".format(assess_id, sm)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'be9e7ddb98b84abe8b9e185b979ccf60','{}');".format(assess_id, pc)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '483a0a1b78264cafbc470ce72d557332','{}');".format(assess_id, eg)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'f9269aebfe2c4d5b9293ba42a40a93ac','{}');".format(assess_id, ta)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '5702908efca4499e87a0239f32920d9b','{}');".format(assess_id, sr)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '4753e55e943c4d418303bf90d599c6b1','{}');".format(assess_id, sa)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'b2af112859d34cada6ce4cf44d393b94','{}');".format(assess_id, sb)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '40d7879025144dbbbf34ba8ea82f060d','{}');".format(assess_id, sd)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'e17d573510904f65a1fe6040b56ad0b1','{}');".format(assess_id, dm)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '53f2da68c37a4ced8d5e767298fba589','{}');".format(assess_id, aa)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '66fb99798fe946e4979a2de98e9d6f8b','{}');".format(assess_id, rt)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'bb5488860c124b6e8076b023485023e1','{}');".format(assess_id, st)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, 'c13aa12c13d04362a3ca3385a8c580ee','{}');".format(assess_id, im)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '53a9cd5c2d3643f3b71e4e9d92b811e2','{}');".format(assess_id, em)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

insert = "INSERT INTO ASSESSPRACTICE VALUES ({}, '8f07145b5ea74388b2217895d5e7b5c2','{}');".format(assess_id, om)
f = open("Database/insert_assessment.pgsql", "a")
f.write(insert + "\n")
f.close()

# Time to do a bunch of iterations to build the SQL to populate the answered question table
# The information needed is in different places to recreate the model and map the answers correctly
# If someone changes the model (or "fixes" it), work will need to be done here to undo the corrections
for row_cells in ws.iter_rows(min_col=1, max_col=7):
    # Get question code from the A column
    quest = row_cells[0].value

    # Get the answer set
    anset_tb = row_cells[-3].value
    
    # I have to look up the answerset letter from the model that doesn't match the answerset letter in the toolbox
    anset = ftoolboxmap.get(anset_tb) 
    
    # Get the answer value
    answ = row_cells[-1].value
    
    # Set the answer id to unknown as an initial value, helped identify that the answersets didn't align
    ansid = 'unknown'

    # If we are on a row that has a question, dive into this
    if(quest != None):
        
        # Iterate through the list of lists model answers looking for a match
        for answer in answers:
            code = answer[0]
            
            # If we find the answer code match between the toolbox question and the model question, proceed
            if code == anset:
                
                # Since there are not individual answer ids, just a group answerset id
                for i in range(3):
                    ansval = answer[1]
                    
                    # If we match the score value in the set of four possible answers for the answer set, collect that UID that we augmented
                    if answ == ansval:
                        ansid = answer[2]

        insert = "INSERT INTO ASSESSEDQUESTION(assessment_id, assessmentquestion_id, answer_id, priority) VALUES ({},'{}','{}',1);".format(assess_id, questions.get(quest),ansid)
        f = open("Database/insert_assessment.pgsql", "a")
        f.write(insert + "\n")
        f.close() 

print('Finished parsing toolbox.')